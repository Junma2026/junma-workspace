"""
比较两张表：2025年进销存-推损益 vs 2025年进销存-推损益_AI
差异用淡红色标出
"""
import openpyxl
from copy import copy

base = r"C:\Users\Administrator\Desktop\药品进销存AI取数"
src1 = rf"{base}\2025年进销存-推损益.xlsx"
src2 = rf"{base}\2025年进销存-推损益_AI.xlsx"
out_path = rf"{base}\2025年进销存-推损益_对比.xlsx"

from openpyxl.styles import PatternFill
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 淡红

wb1 = openpyxl.load_workbook(src1, data_only=True)
ws1 = wb1['Sheet2']
wb2 = openpyxl.load_workbook(src2, data_only=True)
ws2 = wb2['Sheet2']

# 读取源表数据（有格式）
wb1_fmt = openpyxl.load_workbook(src1)
ws1_fmt = wb1_fmt['Sheet2']

# 创建新表
wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = 'Sheet2'

# 复制源表全部内容和格式过来，差异格标红
max_r = ws1_fmt.max_row
max_c = ws1_fmt.max_column

def v(v):
    """获取值用于比较"""
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return round(v, 4)
    return str(v)

# 第1行：大标题
for c in range(1, max_c + 1):
    src_cell = ws1_fmt.cell(1, c)
    tgt_cell = ws_out.cell(1, c)
    tgt_cell.value = src_cell.value
    if src_cell.font:
        tgt_cell.font = copy(src_cell.font)
    if src_cell.alignment:
        tgt_cell.alignment = copy(src_cell.alignment)

# 第2行：表头
for c in range(1, max_c + 1):
    src_cell = ws1_fmt.cell(2, c)
    tgt_cell = ws_out.cell(2, c)
    tgt_cell.value = src_cell.value
    if src_cell.font:
        tgt_cell.font = copy(src_cell.font)
    if src_cell.alignment:
        tgt_cell.alignment = copy(src_cell.alignment)

# 数据行
# 先建立源表2的name->数据映射
data2 = {}
for r in range(3, ws2.max_row + 1):
    name = ws2.cell(r, 2).value
    if name:
        vals = [ws2.cell(r, c).value for c in range(1, 18)]
        data2[str(name).strip()] = vals

# 遍历源表1
changes = 0
keeps = 0
for r in range(3, max_r + 1):
    name1 = ws1_fmt.cell(r, 2).value
    if name1 is None or str(name1).strip() == '':
        continue
    name1 = str(name1).strip()
    
    key = name1
    vals2 = data2.get(key)
    
    for c in range(1, max_c + 1):
        src_cell = ws1_fmt.cell(r, c)
        tgt_cell = ws_out.cell(r, c)
        
        # 复制原始内容
        tgt_cell.value = src_cell.value
        if src_cell.font:
            tgt_cell.font = copy(src_cell.font)
        if src_cell.alignment:
            tgt_cell.alignment = copy(src_cell.alignment)
        if src_cell.number_format:
            tgt_cell.number_format = src_cell.number_format
        
        # 比较数据（仅比较数量列：6-17）
        if vals2 and c >= 6 and c <= 17:
            val1 = v(src_cell.value)
            val2 = v(vals2[c - 1])
            if abs(val1 - val2) > 0.001:
                tgt_cell.fill = red_fill
                changes += 1
        else:
            keeps += 1

# 列宽
for col in ws_out.columns:
    max_len = 0
    letter = col[0].column_letter
    for cell in col:
        try:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        except:
            pass
    ws_out.column_dimensions[letter].width = min(max_len + 4, 40)

wb_out.save(out_path)
print(f"完成: {out_path}")
print(f"差异标记数: {changes} 个单元格")
