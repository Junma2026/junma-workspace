"""
重新生成2025年进销存-推损益表（公式正确版）
公式：期初(24盘点) + 盘点后发药 + 25采购 - 25发药 - 25期末 = 损益数量
"""
import openpyxl
from openpyxl.styles import Font
import os

src = r"C:\Users\Administrator\.openclaw\media\inbound\2025年进销存-推损益---7670e388-dd9a-452c-9efa-228c061d836c.xlsx"
out_dir = r"C:\Users\Administrator\Desktop\药品进销存AI取数"
out_path = os.path.join(out_dir, "2025年进销存-推损益_AI.xlsx")

wb = openpyxl.load_workbook(src, data_only=True)
ws = wb['Sheet2']

# 读入全部数据
data = []
for r in range(3, ws.max_row + 1):
    vals = [ws.cell(r, c).value for c in range(1, 18)]
    code = vals[0]
    name = vals[1]
    spec = vals[2]
    unit = vals[3]
    mfr = vals[4]
    
    # 空行跳过
    if name is None or str(name).strip() == '':
        continue
    
    def n(v):
        """安全转数字"""
        if v is None:
            return 0
        if isinstance(v, (int, float)):
            return v
        try:
            return float(v)
        except:
            return 0
    
    f = n(vals[5])      # 期初数量
    f_amt = n(vals[6])  # 期初总进价
    h = n(vals[7])      # 盘点后发药数量
    h_amt = n(vals[8])  # 盘点后发药总进价
    j = n(vals[9])      # 购药数量
    j_amt = n(vals[10]) # 购药总进价
    l = n(vals[11])     # 发药数量
    l_amt = n(vals[12]) # 发药总进价
    p = n(vals[15])     # 实盘数量
    p_amt = n(vals[16]) # 实盘总进价
    
    # 计算损益数量：期初 + 盘点后发药 + 购药 - 发药 - 实盘
    n_qty = round(f + h + j - l - p, 2)
    
    # 计算损益金额
    if abs(n_qty) > 0.001:
        # 单价：优先用原始表损益金额推算
        orig_n = n(vals[13])
        orig_n_amt = n(vals[14])
        if abs(orig_n) > 0.001:
            unit_price = abs(round(orig_n_amt / orig_n, 4))
        elif h > 0.001 and h_amt > 0:
            unit_price = abs(round(h_amt / h, 4))
        elif j > 0.001 and j_amt > 0:
            unit_price = abs(round(j_amt / j, 4))
        elif l > 0.001 and l_amt > 0:
            unit_price = abs(round(l_amt / l, 4))
        elif p > 0.001 and p_amt > 0:
            unit_price = abs(round(p_amt / p, 4))
        elif f > 0.001 and f_amt > 0:
            unit_price = abs(round(f_amt / f, 4))
        else:
            unit_price = 0
        n_amt = round(n_qty * unit_price, 2)
    else:
        n_amt = 0
    
    data.append({
        'code': code,
        'name': name,
        'spec': spec,
        'unit': unit,
        'mfr': mfr,
        'f': f, 'f_amt': f_amt,
        'h': h, 'h_amt': h_amt,
        'j': j, 'j_amt': j_amt,
        'l': l, 'l_amt': l_amt,
        'n_qty': n_qty, 'n_amt': n_amt,
        'p': p, 'p_amt': p_amt,
    })

# ================================================================
# 创建新Excel
# ================================================================
wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = 'Sheet2'

# 第1行：大标题
titles = ['', '', '', '', '', '2024年期末盘点', '', '2024年期末盘点后的发药', '', '2025年购药', '', '2025年发药', '', '2025年损益', '', '2025年期末', '']
for ci, t in enumerate(titles, 1):
    ws2.cell(1, ci).value = t

# 第2行：列标题
headers = ['编码', '品名', '规格', '基本单位', '生产商',
           '期初数量', '期初总进价', '盘点后发药数量', '盘点后发药总进价',
           '购药数量', '购药总进价', '发药数量', '发药总进价',
           '损益数量', '损益金额', '实盘数量', '实盘总进价']
for ci, h in enumerate(headers, 1):
    ws2.cell(2, ci).value = h

# 数据行
for ri, d in enumerate(data, 3):
    ws2.cell(ri, 1).value = d['code']
    ws2.cell(ri, 2).value = d['name']
    ws2.cell(ri, 3).value = d['spec']
    ws2.cell(ri, 4).value = d['unit']
    ws2.cell(ri, 5).value = d['mfr']
    ws2.cell(ri, 6).value = d['f']
    ws2.cell(ri, 7).value = d['f_amt']
    ws2.cell(ri, 8).value = d['h']
    ws2.cell(ri, 9).value = d['h_amt']
    ws2.cell(ri, 10).value = d['j']
    ws2.cell(ri, 11).value = d['j_amt']
    ws2.cell(ri, 12).value = d['l']
    ws2.cell(ri, 13).value = d['l_amt']
    ws2.cell(ri, 14).value = d['n_qty']
    ws2.cell(ri, 15).value = d['n_amt']
    ws2.cell(ri, 16).value = d['p']
    ws2.cell(ri, 17).value = d['p_amt']

# 列宽
for col in ws2.columns:
    max_len = 0
    letter = col[0].column_letter
    for cell in col:
        try:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        except:
            pass
    ws2.column_dimensions[letter].width = min(max_len + 4, 40)

# ================================================================
# 公式验证
# ================================================================
print("=" * 60)
print("公式：24盘点(期初) + 盘点后发药 + 25购药 - 25发药 - 25期末 = 损益")
print("=" * 60)
ok = bad = 0
for d in data:
    calced = round(d['f'] + d['h'] + d['j'] - d['l'] - d['p'], 2)
    if abs(calced - d['n_qty']) < 0.01:
        ok += 1
    else:
        bad += 1
        print(f"❌ {d['name']:25s} 期初={d['f']:>6}+发药={d['h']:>6}+购药={d['j']:>6}-发药{d['l']:>6}-期末={d['p']:>6}={calced:>8} 损益={d['n_qty']:>8}")

print(f"\n匹配: {ok} 条 ✓")
print(f"不匹配: {bad} 条")

# 统计
print("\n=== 汇总 ===")
loss_items = [d for d in data if d['n_qty'] < 0]
gain_items = [d for d in data if d['n_qty'] > 0]
zero_items = [d for d in data if d['n_qty'] == 0]
print(f"盘亏(负): {len(loss_items)} 种, 金额 {sum(d['n_amt'] for d in loss_items):.2f}")
print(f"盘盈(正): {len(gain_items)} 种, 金额 {sum(d['n_amt'] for d in gain_items):.2f}")
print(f"持平: {len(zero_items)} 种")
print(f"净损益: {sum(d['n_amt'] for d in data):.2f}")

wb2.save(out_path)
print(f"\n完成: {out_path}")
